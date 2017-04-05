import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook

def checkTemplate(templateFileList, maxValue):
    dummyList = []
    for dummy in range(maxValue+1):
        dummyList.append(dummy)

    templateFileList.sort()
    for dummy in dummyList:
        if dummy not in templateFileList:
            return False

    return True

def createTable(filname,templateList,maxValue,inputList):
    if os.path.isfile(filname):
        print "File exists"
        # Create a workbook and add a worksheet.
        wb = Workbook()
        ws1 = wb.active

        row = 1
        col = 1

        inputDict = {}
        for item in range(len(inputList)):
            inputDict[item] = inputList[item]

        if len(inputDict) > (maxValue+1):
            print 'Number of values in the list are greater than template! Aborting...'
            sys.exit(2)

        # Iterate over the data and write it out row by row.
        for itms in templateList:

            rowValue = itms['lineNumber']
            colValue = itms['numberOfItems']
            placesList = itms['lineItems']
            for i in range(colValue):
                if placesList[i] == None:
                    pass
                else:
                    intVal  = int(placesList[i])

                if intVal in inputDict:
                    ws1.cell(row=rowValue,column=i+1,value=inputDict[int(placesList[i])])
                else:
                    ws1.cell(row=rowValue,column=i+1,value=None)


        wb.save(filename = 'temp.xlsx')
        wb1 = load_workbook(filename = filname)
        wb2 = load_workbook(filename = 'temp.xlsx')
        ws1 = wb1.active
        ws2 = wb2.active
        rows = ws2.rows

        insertMarkerRow = ws1.max_row + 2
        colNameList = list()
        for row in rows:
            for cell in row:
                colNameList.append(str(cell.column))

        myset = set(colNameList)
        newColNameList = list(myset)
        newColNameList.sort()

        colDict = {}
        for itm in range(len(newColNameList)):
            colDict[itm] = newColNameList[itm]


        rows = ws2.rows
        for row in rows:
            for cell in row:
                for key in colDict.keys():
                    if colDict[key] == str(cell.column):
                        ws1.cell(row=insertMarkerRow+cell.row,column=key+1,value=cell.value)


        wb1.save(filename = filname)
        wb2.save(filename = 'temp.xlsx')
        os.remove('temp.xlsx')

    else:
        print "The file is missing, new one is created"
        # Create a workbook and add a worksheet.
        wb = Workbook()
        ws1 = wb.active

        row = 1
        col = 1

        inputDict = {}
        for item in range(len(inputList)):
            inputDict[item] = inputList[item]

        if len(inputDict) > (maxValue+1):
            print 'Number of values in the list are greater than template! Aborting...'
            sys.exit(2)

        # Iterate over the data and write it out row by row.
        for itms in templateList:
            rowValue = itms['lineNumber']
            colValue = itms['numberOfItems']
            placesList = itms['lineItems']
            for i in range(colValue):
                if placesList[i] == None:
                    pass
                else:
                    intVal  = int(placesList[i])
                if intVal in inputDict:
                    ws1.cell(row=rowValue,column=i+1,value=inputDict[int(placesList[i])])
                else:
                    ws1.cell(row=rowValue,column=i+1,value=None)

        wb.save(filename = filname)

def main(argv):
    if len(argv) == 2:
        filname = argv[1]
        # Some data we want to write to the worksheet.
        inputList = list(raw_input("Enter the list of items:\n").split(','))
        createTable(filname,inputList)
        sys.exit(0)
    elif len(argv) > 2:
        filesList = argv[1:]
        templateFile = str(filesList[len(filesList)-2])
        outFile = str(filesList[len(filesList)-1])
        filesList.remove(outFile)
        filesList.remove(templateFile)

        templateLines = []
        lineCount = 0
        try:
            wbTemp = load_workbook(filename = templateFile)
            wsTemp = wbTemp.active
            rowsTemp = wsTemp.rows
            valueList = []
            row_count = 0

            for rowTemp in rowsTemp:
                lineDetails = {}
                col_count = 0
                row_count += 1
                cell_items = []
                for cellTemp in rowTemp:
                    col_count += 1
                    cell_items.append(cellTemp.value)

                    if cellTemp.value == None:
                        pass
                    else:
                        valueList.append(int(cellTemp.value))

                lineDetails['lineNumber'] = row_count
                lineDetails['numberOfItems'] = col_count
                lineDetails['lineItems'] = cell_items
                templateLines.append(lineDetails)

            valueList.sort()

            maxValue = max(valueList)

            vaildTemplate = checkTemplate(valueList,maxValue)
            if not vaildTemplate:
                print 'Not a valid Template, missing values..!Abort..'
                sys.exit(2)

        except IOError:
            print '%s does not exist...!Aborting process'%templateFile
            sys.exit(2)

        for item in filesList:
            lines = []
            try:
                with open(item) as file:
                    for line in file:
                        line = line.strip('\n').split(',') #or someother preprocessing
                        lines = lines+line

                createTable(outFile,templateLines,maxValue,lines)
            except IOError:
                print '%s does not exist...!Aborting process'%item
                sys.exit(2)


        print 'Process Successful...'
        sys.exit(0)
    else:
        print 'Atleast One command-line argument is needed:'
        print('Usage: %s [destinationFile]' % argv[0])
        sys.exit(2)

if __name__ == '__main__':
    main(sys.argv)
